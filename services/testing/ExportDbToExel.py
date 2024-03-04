import db.database as db
from openpyxl import Workbook

def create_exel(outname: str):
    wb = Workbook()
    ws = wb.active

    list_users = db.get_all_verify_users()

    index = 1
    ws.title = "Ваші контакти"
    ws.cell(row=index, column=1, value="Ім'я")
    ws.cell(row=index, column=2, value="Нік у телеграмі")
    ws.cell(row=index, column=3, value="Номер телефону")
    for i in list_users:
        index += 1
        ws.cell(row=index, column=1, value=i.name)
        ws.cell(row=index, column=2, value=i.tg_name)
        ws.cell(row=index, column=3, value=i.phone)

    ws2 = wb.create_sheet(title="З інших чатів")

    list_users = db.get_user_other_order()
    index = 1
    ws2.cell(row=index, column=1, value="Ім'я")
    ws2.cell(row=index, column=2, value="Нік у телеграмі")
    ws2.cell(row=index, column=3, value="Телефон")
    ws2.cell(row=index, column=4, value="З якого чату")
    for i in list_users:
        index += 1
        ws2.cell(row=index, column=1, value=i.name)
        ws2.cell(row=index, column=2, value=i.tg_name)
        ws2.cell(row=index, column=3, value=i.phone)
        ws2.cell(row=index, column=4, value=i.fromchanel)

    wb.save(outname + ".xlsx")
    return outname + ".xlsx"